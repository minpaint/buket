# -*- coding: utf-8 -*-
from django.core.management.base import BaseCommand
import openpyxl
from shop.models import Product, Category


class Command(BaseCommand):
    help = 'Assign categories to products based on Excel file'

    def handle(self, *args, **options):
        cats = {c.id: c for c in Category.objects.all()}
        cat_names = {c.name.lower().strip(): c for c in Category.objects.all()}

        self.stdout.write("=== CATEGORIES ===")
        for cid, c in sorted(cats.items()):
            self.stdout.write(f"  {cid}: {c.name}")

        # Load Excel
        wb = openpyxl.load_workbook('G:/Мой диск/buket/bukety_kategorii.xlsx')
        ws = wb.active

        # Build article -> excel_row mapping
        excel_by_art = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            row_id, article, name, composition, price_str, count = row
            if not name:
                continue
            art = str(article).strip() if article else ''
            excel_by_art[art] = {
                'name': name.strip(),
                'composition': str(composition).lower() if composition else '',
            }

        # Also normalized
        excel_norm = {}
        for art, data in excel_by_art.items():
            if art:
                norm = art.lstrip('0') or '0'
                excel_norm[norm] = data
                excel_norm[art] = data

        def find_cat(keyword):
            """Find a category whose name contains the keyword."""
            for cname, cobj in cat_names.items():
                if keyword in cname:
                    return cobj
            return None

        def classify(name, composition):
            """Determine the best category based on product name and composition."""
            n = name.lower()
            c = composition.lower() if composition else ''

            # Мыло ручной работы
            if 'мыло ручной' in n:
                return find_cat('мыло')

            # Мыло букет
            if 'мыло' in n and 'букет' in n:
                return find_cat('мыло')
            if 'мыло' in n:
                return find_cat('мыло')

            # Игрушка из цветов
            if 'игрушк' in n:
                return find_cat('игрушк')

            # Корзина
            if 'корзин' in n:
                return find_cat('корзин')

            # Композиция в коробке
            if 'композиц' in n and 'коробк' in n:
                return find_cat('коробк')
            if 'в коробке' in n:
                return find_cat('коробк')

            # Композиция в корзине
            if 'композиц' in n and 'корзин' in n:
                return find_cat('корзин')

            # Композиция
            if 'композиц' in n:
                return find_cat('композиц')

            # Букет невесты
            if 'невест' in n:
                return find_cat('невест')

            # Свадебные
            if 'свадеб' in n:
                return find_cat('свадеб')

            # Траурные
            if 'траур' in n or 'похорон' in n or 'ритуал' in n:
                return find_cat('траур') or find_cat('ритуал')

            # Сухоцветы
            if 'сухоцвет' in n:
                return find_cat('сухоцвет')

            # Премиум
            if 'премиум' in n:
                return find_cat('премиум')

            # Интерьер
            if 'интерьер' in n:
                return find_cat('интерьер')

            # Авторский
            if 'авторск' in n:
                return find_cat('авторск')

            # Now check by specific flowers in name
            # Розы
            if 'из роз' in n or 'роз ' in n or n.endswith(' роз'):
                r = find_cat('из роз') or find_cat('роз')
                if r:
                    return r

            # Пионы
            if 'пион' in n:
                return find_cat('пион')

            # Хризантемы
            if 'хризантем' in n:
                return find_cat('хризантем')

            # Тюльпаны
            if 'тюльпан' in n:
                return find_cat('тюльпан')

            # Герберы
            if 'гербер' in n:
                return find_cat('гербер')

            # Лилии
            if 'лили' in n:
                return find_cat('лили')

            # Орхидеи
            if 'орхиде' in n:
                return find_cat('орхиде')

            # Ирисы
            if 'ирис' in n and 'ирис' not in 'кризис':
                return find_cat('ирис')

            # Гвоздики
            if 'гвоздик' in n:
                return find_cat('гвоздик')

            # Альстромерии
            if 'альстромери' in n:
                return find_cat('альстромери')

            # Эустомы
            if 'эустом' in n:
                return find_cat('эустом')

            # Гортензии
            if 'гортензи' in n:
                return find_cat('гортензи')

            # Check composition tags for category hints
            comp_tags = [t.strip() for t in c.split(',')]

            # Мыло ручной работы in composition
            if 'мыло ручной' in c:
                return find_cat('мыло')

            # Авторская композиция
            if any('авторск' in t for t in comp_tags):
                return find_cat('авторск')

            # Композиция в корзине
            if any('корзин' in t for t in comp_tags):
                # But only if the product is a composition
                if 'композиц' in n:
                    return find_cat('корзин')

            # Check for event-based categories
            if '8 март' in n or '8 марта' in c:
                return find_cat('8 март')
            if '23 фев' in n or '23 февраля' in c:
                return find_cat('23 фев')
            if 'день рожд' in n:
                return find_cat('рожд')
            if 'валентин' in n:
                return find_cat('валентин')
            if 'день учител' in n:
                return find_cat('учител')

            # Check for "букет" with specific keywords
            if 'букет' in n:
                # Монобукет (single flower type)
                if 'моно' in n:
                    return find_cat('моно')
                # Букет из полевых
                if 'полев' in n:
                    return find_cat('полев')

                # Classify by main composition
                # "Авторский букет" in composition
                if any('авторск' in t for t in comp_tags):
                    return find_cat('авторск')

                # If has specific tags, try to match
                main_flowers = {
                    'роз': 'роз',
                    'хризантем': 'хризантем',
                    'пион': 'пион',
                    'тюльпан': 'тюльпан',
                    'гербер': 'гербер',
                    'лили': 'лили',
                    'гвоздик': 'гвоздик',
                    'ирис': 'ирис',
                }
                for flower, cat_key in main_flowers.items():
                    if flower in n:
                        r = find_cat(cat_key)
                        if r:
                            return r

            # Комнатные / горшечные
            if 'комнатн' in n or 'горшк' in n or 'горшок' in n or 'горшеч' in n:
                return find_cat('комнатн') or find_cat('горшк')

            # Шар
            if 'шар ' in n or 'шары' in n or n.endswith(' шар'):
                return find_cat('шар')

            # Default - if nothing matched
            # Check if it's a mix bouquet
            if 'букет' in n and 'микс' in n:
                return find_cat('микс')

            # Подарки / сладости
            if 'сладост' in n or 'конфет' in n or 'шоколад' in n:
                return find_cat('сладост') or find_cat('подарк')

            return None

        # Assign categories
        updated = 0
        already_ok = 0
        no_category = 0
        results = []

        for p in Product.objects.all().order_by('id'):
            art = p.article.strip()
            norm = art.lstrip('0') or '0'

            excel = excel_by_art.get(art) or excel_norm.get(norm) or excel_norm.get(art)

            if excel:
                cat_obj = classify(excel['name'], excel['composition'])
                if not cat_obj:
                    cat_obj = classify(p.title, excel['composition'])
            else:
                cat_obj = classify(p.title, '')

            if cat_obj:
                if p.category_id != cat_obj.id:
                    old_cat = p.category.name if p.category else 'None'
                    results.append(f"UPDATE {p.id}|{p.article}|{p.title}: {old_cat} -> {cat_obj.name}")
                    p.category = cat_obj
                    p.save()
                    updated += 1
                else:
                    already_ok += 1
            else:
                no_category += 1
                comp_info = excel['composition'][:80] if excel else ''
                results.append(f"NO_CAT {p.id}|{p.article}|{p.title} [{comp_info}]")

        self.stdout.write("")
        self.stdout.write("=== RESULTS ===")
        for r in results:
            self.stdout.write(f"  {r}")

        self.stdout.write("")
        self.stdout.write(f"Updated: {updated}")
        self.stdout.write(f"Already correct: {already_ok}")
        self.stdout.write(f"Could not determine category: {no_category}")
